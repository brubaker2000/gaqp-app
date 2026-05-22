#!/usr/bin/env python3
"""
Generate GAQP Standards Package xlsx distribution artifact from JSON source files.
JSON is the source of truth. This script produces the xlsx output.

Usage:
    pip install -r requirements.txt
    python generate_xlsx.py [--version v0.1-working] [--out ./dist]
"""

import argparse
import hashlib
import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Missing dependency: pip install openpyxl")


STANDARDS_ROOT = Path(__file__).parent.parent / "standards"

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ALT_FILL = PatternFill(start_color="EEF3FF", end_color="EEF3FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def load_json(version_dir: Path, filename: str) -> dict:
    path = version_dir / filename
    with open(path) as f:
        return json.load(f)


def style_header(cell, sub=False):
    cell.font = SUBHEADER_FONT if sub else HEADER_FONT
    cell.fill = SUBHEADER_FILL if sub else HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = THIN_BORDER


def style_cell(cell, alt=False):
    cell.fill = ALT_FILL if alt else PatternFill()
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = THIN_BORDER


def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_claim_types(wb, data: dict):
    ws = wb.create_sheet("Claim Types")
    ws.freeze_panes = "A2"

    headers = ["#", "claim_type", "Operational Definition"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)

    for i, ct in enumerate(data["claim_types"]):
        row = i + 2
        alt = i % 2 == 1
        for col, val in enumerate([ct["id"], ct["claim_type"], ct["definition"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            style_cell(c, alt)

    note_row = len(data["claim_types"]) + 3
    ws.cell(row=note_row, column=1, value="Register Status:").font = Font(bold=True)
    ws.cell(row=note_row, column=2, value=data["register_status"].upper())
    ws.cell(row=note_row + 1, column=1, value="Note:").font = Font(bold=True)
    c = ws.cell(row=note_row + 1, column=2, value=data["register_note"])
    c.fill = NOTE_FILL
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{note_row + 1}:C{note_row + 1}")

    forbidden_start = note_row + 3
    ws.cell(row=forbidden_start, column=1, value="FORBIDDEN LABELS").font = Font(bold=True, color="C00000")
    ws.cell(row=forbidden_start + 1, column=1, value="Forbidden Label")
    ws.cell(row=forbidden_start + 1, column=2, value="Remap To")
    ws.cell(row=forbidden_start + 1, column=3, value="Note")
    style_header(ws.cell(row=forbidden_start + 1, column=1), sub=True)
    style_header(ws.cell(row=forbidden_start + 1, column=2), sub=True)
    style_header(ws.cell(row=forbidden_start + 1, column=3), sub=True)

    for i, fl in enumerate(data["forbidden_labels"]):
        r = forbidden_start + 2 + i
        ws.cell(row=r, column=1, value=fl["forbidden"])
        ws.cell(row=r, column=2, value=" / ".join(fl["remap_to"]))
        ws.cell(row=r, column=3, value=fl.get("note", ""))

    set_col_widths(ws, [6, 30, 70])


def write_admission_tests(wb, data: dict):
    ws = wb.create_sheet("Admission Tests")
    ws.freeze_panes = "A2"

    headers = ["#", "Test", "Pass Condition"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, t in enumerate(data["tests"]):
        row = i + 2
        alt = i % 2 == 1
        for col, val in enumerate([t["id"], t["name"], t["pass_condition"]], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    note_row = len(data["tests"]) + 3
    ws.cell(row=note_row, column=1, value="Operational Discipline:").font = Font(bold=True)
    c = ws.cell(row=note_row, column=2, value=data["operational_discipline"])
    c.fill = NOTE_FILL
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"B{note_row}:C{note_row}")

    set_col_widths(ws, [6, 25, 80])


def write_confidence_ladder(wb, data: dict):
    ws = wb.create_sheet("Confidence Ladder")
    ws.freeze_panes = "A2"

    headers = ["Level", "Score", "Condition"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, lvl in enumerate(data["levels"]):
        row = i + 2
        alt = i % 2 == 1
        score = lvl["score"] if lvl["score"] is not None else "frozen"
        for col, val in enumerate([lvl["level"], score, lvl["condition"]], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    note_row = len(data["levels"]) + 3
    for label, key in [("Promotion Rule:", "promotion_rule"), ("Independence:", "independence_definition")]:
        ws.cell(row=note_row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=note_row, column=2, value=data[key])
        c.fill = NOTE_FILL
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{note_row}:C{note_row}")
        note_row += 1

    set_col_widths(ws, [18, 10, 75])


def write_metadata_schema(wb, data: dict):
    ws = wb.create_sheet("Metadata Schema")
    ws.freeze_panes = "A2"

    headers = ["Field", "Group", "Required", "Type", "Definition"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, f in enumerate(data["fields"]):
        row = i + 2
        alt = i % 2 == 1
        req = "Required" if f.get("required") else "Optional"
        definition = f["definition"]
        if "enum" in f:
            definition += f"  [enum: {', '.join(str(v) for v in f['enum'])}]"
        if f.get("nullable"):
            definition += "  [nullable]"
        if "ref" in f:
            definition += f"  [ref: {f['ref']}]"
        for col, val in enumerate([f["field"], f.get("group", ""), req, f.get("type", ""), definition], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    removed_start = len(data["fields"]) + 3
    ws.cell(row=removed_start, column=1, value="REMOVED FIELDS").font = Font(bold=True, color="C00000")
    headers2 = ["Field", "Reason"]
    for col, h in enumerate(headers2, 1):
        style_header(ws.cell(row=removed_start + 1, column=col), sub=True)
        ws.cell(row=removed_start + 1, column=col, value=h)
    for i, rf in enumerate(data.get("removed_fields", [])):
        r = removed_start + 2 + i
        ws.cell(row=r, column=1, value=rf["field"])
        ws.cell(row=r, column=2, value=rf["reason"])

    set_col_widths(ws, [30, 18, 12, 12, 70])


def write_source_anchor_schema(wb, data: dict):
    ws = wb.create_sheet("Source Anchor Schema")
    ws.freeze_panes = "A2"

    headers = ["Field", "Notes"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, f in enumerate(data["fields"]):
        row = i + 2
        alt = i % 2 == 1
        notes = f["notes"]
        if "enum" in f:
            notes += f" [enum: {', '.join(f['enum'])}]"
        for col, val in enumerate([f["field"], notes], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    set_col_widths(ws, [25, 70])


def write_conformance_levels(wb, data: dict):
    ws = wb.create_sheet("Conformance Levels")
    ws.freeze_panes = "A2"

    headers = ["Level", "Name", "Capability"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, lvl in enumerate(data["levels"]):
        row = i + 2
        alt = i % 2 == 1
        for col, val in enumerate([lvl["level"], lvl["name"], lvl["capability"]], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    set_col_widths(ws, [10, 28, 70])


def write_sector_vocabulary(wb, data: dict):
    ws = wb.create_sheet("Sector Vocabulary")
    ws.freeze_panes = "A2"

    headers = ["Code", "Sector Name", "GICS Sector", "SIC Range", "NAICS Range", "Mapping Notes"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    sorted_sectors = sorted(data["sectors"], key=lambda s: s["sector_name"])
    for i, sector in enumerate(sorted_sectors):
        row = i + 2
        alt = i % 2 == 1
        for col, val in enumerate([
            sector["sector_code"],
            sector["sector_name"],
            sector.get("gics_sector", ""),
            sector.get("sic_range", ""),
            sector.get("naics_range", ""),
            sector.get("mapping_notes", ""),
        ], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    fallback = data.get("fallback")
    if fallback:
        r = len(sorted_sectors) + 3
        ws.cell(row=r, column=1, value="Fallback:").font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=fallback)
        c.fill = NOTE_FILL

    set_col_widths(ws, [8, 28, 22, 18, 18, 50])


def write_principles(wb, data: dict):
    ws = wb.create_sheet("Principles")
    ws.freeze_panes = "A2"

    headers = ["#", "Principle", "Requirement"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    for i, p in enumerate(data["principles"]):
        row = i + 2
        alt = i % 2 == 1
        for col, val in enumerate([p["id"], p["name"], p["requirement"]], 1):
            style_cell(ws.cell(row=row, column=col, value=val), alt)

    set_col_widths(ws, [6, 35, 75])


def write_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    def heading(row, text, size=13, color="1F3864"):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(size=size, bold=True, color=color)
        ws.merge_cells(f"A{row}:B{row}")
        c.alignment = Alignment(wrap_text=True)
        return row + 1

    def body(row, text, bold_label=None):
        if bold_label:
            ws.cell(row=row, column=1, value=bold_label).font = Font(bold=True, color="2F5496")
        c = ws.cell(row=row, column=2 if bold_label else 1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not bold_label:
            ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 30
        return row + 1

    def spacer(row):
        ws.row_dimensions[row].height = 8
        return row + 1

    r = 1
    c = ws.cell(row=r, column=1, value="GAQP Standards Package — Deconstructor's Guide")
    c.font = Font(size=16, bold=True, color="1F3864")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 28
    r += 1
    r = spacer(r)

    r = heading(r, "What this workbook is")
    r = body(r, (
        "This workbook is the GAQP Standards Package — the machine-readable rulebook for "
        "GAQP-compliant document deconstruction. It defines every rule a deconstructor must "
        "follow: what counts as a valid claim, how to classify it, how to tag it, and how to "
        "package it in a .intel sidecar file. All sheets are derived from the JSON source files "
        "in the IQS standards repository."
    ))
    r = spacer(r)

    r = heading(r, "What deconstruction is — and is not")
    r = body(r, (
        "Deconstruction is the governed extraction of atomic qualitative claims from a source document. "
        "Each extracted claim (a 'nugget') must pass all seven admission tests before it is admitted. "
        "Current working output: a spreadsheet of admitted nuggets. Future output: a .intel sidecar file "
        "that travels with the source document. What happens to nuggets after delivery — querying, "
        "synthesis, corroboration, reporting — is reconstruction, which is outside GAQP's jurisdiction."
    ))
    r = spacer(r)

    r = heading(r, "Deconstruction workflow")
    steps = [
        ("1. Read the source document", "Identify the document type, creation date, author, and page structure. Record these in the .intel source_binding section."),
        ("2. Apply the seven admission tests", "See the Admission Tests sheet. Every candidate claim must pass all seven gates. A claim that fails any test is not born — it is not recorded anywhere."),
        ("3. Assign a claim type", "See the Claim Types sheet. Choose exactly one type from the closed 25-type register. Use the forbidden labels remap table if tempted to use an unlisted label."),
        ("4. Write the source quote", "Copy the verbatim passage from the source document that the claim is drawn from. This is required. Paraphrase is not permitted."),
        ("5. Set the inference flag", "If the claim goes beyond what the source explicitly states, set inference_flag = TRUE. Never hide inference."),
        ("6. Assign sector and domain tags", "See Sector Vocabulary and Domain Subtags sheets. Assign primary_sector and domain_tags_pipe. NULL is valid for sector-agnostic claims."),
        ("7. Assign cross-cutting tags", "See Tags Vocabulary sheet. Assign tags_pipe using func:, risk:, topic:, and geo: prefixes. NULL is valid."),
        ("8. Set confidence level", "See Confidence Ladder sheet. New claims extracted from a single source start at 'seed' (0.50). Confidence advances only through independent corroboration."),
        ("9. Deliver the output", (
            "Current version: export all admitted nuggets to a spreadsheet. Each row is one admitted claim "
            "carrying all 29 fields. Multiple source documents can be deconstructed and their nuggets combined "
            "into a single spreadsheet for cross-document analysis. "
            "Future version: package admitted claims into a .intel sidecar file that travels with the source document. "
            "The .intel spec is defined in this package — implementation is in progress."
        )),
    ]
    for label, text in steps:
        r = body(r, text, bold_label=label)
    r = spacer(r)

    r = heading(r, "Key rules — do not violate these")
    rules = [
        ("Source locality", "A .intel file contains only claims from its corresponding source document. No cross-document content."),
        ("Existence = admitted", "If a claim record is in the .intel file, it passed all seven tests. There are no rejected or needs-review records in a compliant .intel file."),
        ("Source quote required", "Every admitted claim must carry the verbatim source_quote. No exceptions."),
        ("Inference must be labeled", "inference_flag = TRUE for any claim that goes beyond the source text. Never hide inference."),
        ("Claim types are closed", "The 25-type register is the complete list. Do not invent new types. Use the forbidden labels remap table."),
        ("Confidence starts at seed", "A single-source extraction starts at seed (0.50). It cannot be self-promoted — only independent corroboration advances confidence."),
        ("Security inherits upward", "A .intel sidecar must carry access controls equal to or stricter than its source document. Never weaker."),
    ]
    for label, text in rules:
        r = body(r, text, bold_label=label)
    r = spacer(r)

    r = heading(r, "Sheet navigation guide")
    nav = [
        ("Cover", "Package identity, version, status, and checksum."),
        ("Claim Types", "The closed 25-type claim register with definitions and forbidden label remaps."),
        ("Admission Tests", "The seven gates every candidate claim must pass."),
        ("Confidence Ladder", "Six confidence levels, numeric scores, and promotion rules."),
        ("Metadata Schema", "The 29 required and optional fields every admitted claim record must carry."),
        ("Source Anchor Schema", "The 20 fields in a source anchor record — exact location within the source document."),
        ("Conformance Levels", "L1 through L7 — the compliance maturity ladder for deconstruction implementations."),
        ("Sector Vocabulary", "30 industry sectors with GICS, SIC, and NAICS crosswalks. Used for primary_sector field."),
        ("Domain Subtags", "3,580 subtags across all 30 sectors organized by dimension. Used for domain_tags_pipe field."),
        ("Tags Vocabulary", "Cross-cutting tag prefixes: func, risk, topic, geo. Used for tags_pipe field."),
        ("Principles", "The 10 constitutional principles governing GAQP."),
    ]
    for label, text in nav:
        r = body(r, text, bold_label=label)


def write_cover(wb, manifest: dict):
    ws = wb.create_sheet("Cover", 1)

    title_cell = ws.cell(row=1, column=1, value=manifest["standard_name"])
    title_cell.font = Font(size=18, bold=True, color="1F3864")
    ws.merge_cells("A1:C1")

    rows = [
        ("Abbreviation", manifest["standard_abbreviation"]),
        ("Standards Body", manifest["standards_body"]),
        ("Package Version", manifest["version"]),
        ("Status", manifest["status"].upper()),
        ("Release Date", manifest["release_date"] or "TBD"),
        ("Checksum (SHA-256)", manifest["checksum_sha256"] or "TBD — set at version freeze"),
        ("Description", manifest["description"]),
    ]

    for i, (label, value) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=value)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{i}:C{i}")

    for label, value in manifest["notes"].items():
        i += 1
        ws.cell(row=i, column=1, value=f"Note — {label}").font = Font(bold=True, color="7F6000")
        c = ws.cell(row=i, column=2, value=value)
        c.fill = NOTE_FILL
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{i}:C{i}")

    set_col_widths(ws, [28, 55, 20])


def write_domain_modules(wb, data: dict):
    ws = wb.create_sheet("Domain Subtags")
    ws.freeze_panes = "A2"

    headers = ["Sector Code", "Sector Name", "Dimension", "Subtag Code", "Subtag Name"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    row = 2
    for sector in data["sectors"]:
        for dim in sector["dimensions"]:
            for i, subtag in enumerate(dim["subtags"]):
                alt = row % 2 == 0
                for col, val in enumerate([
                    sector["sector_code"],
                    sector["sector_name"],
                    dim["dimension"],
                    subtag["code"],
                    subtag["name"],
                ], 1):
                    style_cell(ws.cell(row=row, column=col, value=val), alt)
                row += 1

    note_row = row + 1
    ws.cell(row=note_row, column=1, value=f"Total subtags: {row - 2}").font = Font(bold=True)
    ws.cell(row=note_row, column=2, value=data.get("tagging_rule", "")).fill = NOTE_FILL
    ws.merge_cells(f"B{note_row}:E{note_row}")

    set_col_widths(ws, [12, 28, 30, 20, 40])


def write_tags_vocabulary(wb, data: dict):
    ws = wb.create_sheet("Tags Vocabulary")
    ws.freeze_panes = "A2"

    headers = ["Prefix", "Label", "Value", "Description"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h))

    row = 2
    for prefix_entry in data["prefixes"]:
        for i, value in enumerate(prefix_entry["values"]):
            alt = row % 2 == 0
            for col, val in enumerate([
                prefix_entry["prefix"],
                prefix_entry["label"],
                value,
                prefix_entry["description"] if i == 0 else "",
            ], 1):
                style_cell(ws.cell(row=row, column=col, value=val), alt)
            row += 1

    note_row = row + 1
    ws.cell(row=note_row, column=1, value="Example:").font = Font(bold=True)
    c = ws.cell(row=note_row, column=2, value=data.get("example", ""))
    c.fill = NOTE_FILL
    ws.merge_cells(f"B{note_row}:D{note_row}")
    note_row += 1
    ws.cell(row=note_row, column=1, value="Null rule:").font = Font(bold=True)
    c = ws.cell(row=note_row, column=2, value=data.get("null_rule", ""))
    c.fill = NOTE_FILL
    ws.merge_cells(f"B{note_row}:D{note_row}")

    set_col_widths(ws, [10, 22, 25, 60])


def write_llm_prompt(wb):
    ws = wb.create_sheet("LLM Prompt")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 85

    def heading(row, text, color="1F3864"):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(size=13, bold=True, color=color)
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 20
        return row + 1

    def label_row(row, label, value, note_color=False):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="2F5496")
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if note_color:
            c.fill = NOTE_FILL
        ws.row_dimensions[row].height = 42
        return row + 1

    def spacer(row):
        ws.row_dimensions[row].height = 8
        return row + 1

    r = 1
    c = ws.cell(row=r, column=1, value="GAQP Deconstruction Prompt — Copy and send to your LLM with the source document")
    c.font = Font(size=14, bold=True, color="1F3864")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 24
    r += 1
    r = spacer(r)

    r = heading(r, "How to use this sheet")
    instructions = (
        "1. Copy the prompt block below.\n"
        "2. Open your LLM (Claude, GPT-4, Gemini, etc.).\n"
        "3. Paste the prompt, then attach or paste the source document.\n"
        "4. The LLM will return a CSV table of admitted nuggets.\n"
        "5. Paste the CSV output into a new Excel sheet. Each row is one admitted claim."
    )
    c = ws.cell(row=r, column=1, value=instructions)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 90
    r += 1
    r = spacer(r)

    r = heading(r, "Prompt block — copy everything below this line")

    prompt_text = (
        "You are a GAQP-compliant document deconstructor operating under the Generally Accepted "
        "Qualitative Principles (GAQP) standard, version 0.1-working.\n\n"

        "Your task: extract all admitted qualitative claims (nuggets) from the attached source document. "
        "Apply all seven admission tests to every candidate claim. Only admitted claims appear in your output. "
        "Do not record rejected candidates.\n\n"

        "SEVEN ADMISSION TESTS — every claim must pass all seven:\n"
        "1. Standalone: the claim makes sense without additional context.\n"
        "2. Disputability: a reasonable person could disagree with it.\n"
        "3. Governance: it is a qualitative assertion, not a raw data point.\n"
        "4. Activation: it has operational relevance — it could inform a decision.\n"
        "5. Durability: it will remain meaningful beyond the immediate moment.\n"
        "6. Composability: it can combine with other claims to produce insight.\n"
        "7. Non-triviality: it is not obvious, tautological, or content-free.\n\n"

        "OUTPUT FORMAT — respond with a CSV table only. No prose before or after. "
        "First row is the header. Each subsequent row is one admitted claim.\n\n"

        "CSV COLUMNS (in this exact order):\n"
        "claim_id, claim_text, claim_type, source_quote, inference_flag, "
        "source_page, source_section, source_paragraph, "
        "primary_sector, domain_tags_pipe, tags_pipe, entities_pipe, "
        "confidence_level, confidence_score, nugget_status, "
        "extraction_method, deconstruction_profile, standards_package_version, "
        "deconstructor_note\n\n"

        "FIELD RULES:\n"
        "claim_id: sequential integer starting at 1.\n"
        "claim_text: the governed atomic claim in your own governed wording.\n"
        "claim_type: exactly one type from this closed list — "
        "Axiom, Definition, Ontological Assertion, Principle, Doctrine, Heuristic, Best Practice, "
        "Tendency, Observation, Event, Institutional Precedent, Constraint, Threshold Condition, "
        "Objective, Tradeoff, Causal Claim, Declaration of Value, Diagnostic Signal, "
        "Stakeholder Complaint Claim, Strength, Weakness, Opportunity, Threat, Asset, Liability.\n"
        "source_quote: verbatim passage from the source document. Required. No paraphrase.\n"
        "inference_flag: TRUE if the claim goes beyond what the source explicitly states. FALSE otherwise.\n"
        "source_page: page number or NULL.\n"
        "source_section: section heading or NULL.\n"
        "source_paragraph: paragraph number or NULL.\n"
        "primary_sector: single dominant industry sector or NULL for sector-agnostic claims.\n"
        "domain_tags_pipe: pipe-delimited SectorCode:SubtagCode pairs or NULL. Example: FIN:PE|FIN:LBO\n"
        "tags_pipe: pipe-delimited prefixed tags or NULL. "
        "Prefixes: func: (business function), risk: (risk category), topic: (subject), geo: (geography). "
        "Example: func:Finance|risk:Leverage|topic:Valuation\n"
        "entities_pipe: pipe-delimited named entities or NULL.\n"
        "confidence_level: always 'seed' for new single-source extractions.\n"
        "confidence_score: always 0.50 for seed.\n"
        "nugget_status: always 'active'.\n"
        "extraction_method: always 'llm'.\n"
        "deconstruction_profile: enter the name or identifier of the instruction set you are using, or 'default'.\n"
        "standards_package_version: always '0.1-working'.\n"
        "deconstructor_note: any operator note about this claim — unusual context, ambiguous classification, "
        "notable inference. NULL if none.\n\n"

        "Begin. Output the CSV table only."
    )

    c = ws.cell(row=r, column=1, value=prompt_text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill(start_color="EEF3FF", end_color="EEF3FF", fill_type="solid")
    c.font = Font(name="Courier New", size=9)
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 420
    r += 1
    r = spacer(r)

    r = heading(r, "Fields omitted from LLM output — filled in post-processing")
    omitted = [
        ("fingerprint", "SHA-256 hash of claim_text + source_anchor_id. Computed programmatically after extraction."),
        ("record_type", "Always 'claim'. Added programmatically."),
        ("extracted_at", "ISO 8601 timestamp. Set programmatically at extraction time."),
        ("source_id", "Foreign key to source_binding record. Set programmatically."),
        ("source_document", "Source filename. Set programmatically."),
        ("source_hash", "SHA-256 of source file. Computed programmatically."),
        ("source_anchor_id", "Foreign key to anchor record. Set programmatically."),
        ("relationship_hints_pipe", "Populated during corpus review, not at extraction time."),
        ("validation_flags_pipe", "Populated by validator after extraction."),
        ("review_notes", "Post-admission operator notes. NULL at extraction time."),
    ]
    r = label_row(r, "Field", "Why omitted from LLM prompt")
    for field, reason in omitted:
        r = label_row(r, field, reason)


def compute_package_checksum(version_dir: Path) -> str:
    h = hashlib.sha256()
    for f in sorted(version_dir.glob("*.json")):
        if f.name == "manifest.json":
            continue
        h.update(f.read_bytes())
    return h.hexdigest()


def main():
    parser = argparse.ArgumentParser(description="Generate GAQP Standards Package xlsx from JSON source.")
    parser.add_argument("--version", default="v0.1-working", help="Standards package version directory name")
    parser.add_argument("--out", default="./dist", help="Output directory for xlsx file")
    args = parser.parse_args()

    version_dir = STANDARDS_ROOT / args.version
    if not version_dir.exists():
        raise SystemExit(f"Version directory not found: {version_dir}")

    manifest = load_json(version_dir, "manifest.json")
    claim_types = load_json(version_dir, "claim_types.json")
    admission_tests = load_json(version_dir, "admission_tests.json")
    confidence_ladder = load_json(version_dir, "confidence_ladder.json")
    metadata_schema = load_json(version_dir, "metadata_schema.json")
    source_anchor = load_json(version_dir, "source_anchor_schema.json")
    conformance = load_json(version_dir, "conformance_levels.json")
    sectors = load_json(version_dir, "sector_vocabulary.json")
    principles = load_json(version_dir, "principles.json")
    domain_modules = load_json(version_dir, "domain_modules.json")
    tags_vocabulary = load_json(version_dir, "tags_vocabulary.json")

    checksum = compute_package_checksum(version_dir)
    manifest["checksum_sha256"] = checksum

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_readme(wb)
    write_llm_prompt(wb)
    write_cover(wb, manifest)
    write_claim_types(wb, claim_types)
    write_admission_tests(wb, admission_tests)
    write_confidence_ladder(wb, confidence_ladder)
    write_metadata_schema(wb, metadata_schema)
    write_source_anchor_schema(wb, source_anchor)
    write_conformance_levels(wb, conformance)
    write_sector_vocabulary(wb, sectors)
    write_domain_modules(wb, domain_modules)
    write_tags_vocabulary(wb, tags_vocabulary)
    write_principles(wb, principles)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    version_slug = args.version.replace("/", "-")
    out_path = out_dir / f"GAQP_Standards_Package_{version_slug}.xlsx"
    wb.save(out_path)

    print(f"Generated: {out_path}")
    print(f"Package checksum (SHA-256): {checksum}")


if __name__ == "__main__":
    main()
